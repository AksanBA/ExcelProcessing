#!/usr/bin/env excelprocess
"""
This is a slightly modified code from: https://github.com/elizabethsiegle/nba-stats-twilio-sms-bot
Currently have the following trial number set up for use to return the player comparison: +1213-437-9533
No permanent ngrok webhook at the moment!

Here we expand the project with pandas with the following aims:
1) Given a player and season, return per game statistics
2) Given a set of player and season per game statistics, return the best or worst player in that set (comparison)
3) Use openpyxl to do something non-trivial that cannot be completed with pandas: charts!

"""
from flask import Flask, request
from twilio.twiml.messaging_response import MessagingResponse
from openpyxl import load_workbook, Workbook
import pandas as pd

def parse_data_into_dict(data):
    list_of_players = []
    list_of_stats = []
    stat_dict = {
        "age":"B", "gp":"C","w":"D","l":"E","min":"F","pts":"G",
        "fgm":"H","fga":"I","fg%":"J",
        "3pm":"K","3pa":"L","ftm":"M","fta":"N",
        "ft%":"O","oreb":"P","dreb":"Q","reb":"R","ast":"S","tov":"T", "stl": "U",
        "blk": "V", "pf": "W", "dd2": "X", "td3": "Y"
    }
    excelfile = 'nbastats.xlsx'
    wb = load_workbook(excelfile)
    ws = wb[wb.sheetnames[0]]
    for row in range(1, ws.max_row+1): #need +1 to get last row!
        for col in "A": #A gets players for texted season
            cell_name="{}{}".format(col, row)
            list_of_players.append(ws[cell_name].value.lower())
            for col in stat_dict[data]: # gets column of whatever statistic
                cell_name="{}{}".format(col, row)
                #print(ws[cell_name].value)
                list_of_stats.append(ws[cell_name].value)
    return dict(zip(list_of_players, list_of_stats))

def user_choice():
    season =

def load_year():
    year = input('Enter the season you would like to compare: ')
    pd.read_excel()


app = Flask(__name__)


@app.route('/sms', methods=['GET', 'POST'])
def send_sms():
    msg = request.form['Body'].lower()  # convert to lowercase
    typomsg = "send 1st + last names of 2 players followed by a stat (GP,W,L,MIN,PTS,FG%,3P%,FT%,REB,AST,STL,BLK). Check for typos!"
    player_and_stat = msg.split()  # split

    if len(player_and_stat) == 5:  # check input: 2 players + stat
        player1 = player_and_stat[0] + " " + player_and_stat[1]
        player2 = player_and_stat[2] + " " + player_and_stat[3]
        stat = player_and_stat[4]
        player_stat_map = parse_data_into_dict(stat)
        if player1 in player_stat_map.keys() and player2 in player_stat_map.keys():
            if player_stat_map[player1] > player_stat_map[player2]:
                ret = MessagingResponse()
                ret.message(
                    player1 + " 's total " + str(player_stat_map[player1]) + ", higher than " + player2 + "\'s " + str(
                        player_stat_map[player2]))
            else:
                ret = MessagingResponse()
                ret.message(
                    player2 + " 's total " + str(player_stat_map[player2]) + ", higher than " + player1 + "\'s " + str(
                        player_stat_map[player1]))
        else:  # check
            ret = MessagingResponse()
            ret.message("check both players' names (first and last!)")
    else:  # idk how many players
        ret = MessagingResponse()
        ret.message(typomsg)
    return str(ret)


if __name__ == "__main__":
    app.run(debug=True)
